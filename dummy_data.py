import openpyxl
from faker import Faker

# Create a Faker instance
fake = Faker()

# Open the Excel workbook (change the path to your Excel file)
workbook = openpyxl.load_workbook('E:/python notes/tyk/dummy_list.xlsx')
worksheet = workbook.active

# Define the number of rows of dummy data you want to generate
num_rows = 10  # Change this as needed

# Insert dummy data into Excel
for row in range(2, num_rows + 2):  # Start from the 2nd row (assuming headers in the 1st row)
    worksheet.cell(row=row, column=1, value=fake.first_name())
    worksheet.cell(row=row, column=2, value=fake.last_name())
    worksheet.cell(row=row, column=3, value=fake.email())


# Save the modified workbook
workbook.save('E:/python notes/tyk/dummy_list.xlsx')
